import json
import sys
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
CONFIG_PATH = Path(__file__).resolve().parent / "fixture_runs.json"


def main():
    if len(sys.argv) > 1:
        target = Path(sys.argv[1])
        if not target.is_absolute():
            target = BASE_DIR / target
    else:
        if not CONFIG_PATH.exists():
            raise FileNotFoundError(f"fixture_runs.json not found: {CONFIG_PATH}")
        with CONFIG_PATH.open("r", encoding="utf-8") as fh:
            config = json.load(fh)
        pension_runs = config.get("pension_runs") or []
        if not pension_runs:
            print("No pension_runs defined in fixture_runs.json")
            return
        first_run = pension_runs[0]
        output_file = first_run.get("output_file")
        if not output_file:
            raise ValueError("First excel run has no output_file")
        target = BASE_DIR / output_file

    if not target.exists():
        raise FileNotFoundError(f"File not found: {target}")

    wb = openpyxl.load_workbook(target)
    print(f"File: {target}")
    print(f"Sheets: {wb.sheetnames}")
    print()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        print(f"Sheet: {sheet_name!r}  ({ws.max_row - 1} data rows)")
        if rows:
            header = rows[0]
            for i, cell in enumerate(header):
                print(f"  col {i}: {cell!r}")
        print()
        preview = rows[1:6]
        for row_num, row in enumerate(preview, start=1):
            print(f"  row {row_num}: {row}")
        if ws.max_row - 1 > len(preview):
            print(f"  ... ({ws.max_row - 1 - len(preview)} more rows)")
        print()


if __name__ == "__main__":
    main()
